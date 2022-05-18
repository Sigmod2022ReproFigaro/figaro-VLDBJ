""""
Figaro times collection times script.
"""
import csv
import numpy as np
import pandas as pd
import argparse
import os
from openpyxl import load_workbook, Workbook
from openpyxl.utils.cell import get_column_letter
import logging
from data_management.database import Database
from data_management.database_psql import DatabasePsql
from data_management.query import Query
import argparse
import sys
import numpy as np
from evaluation.custom_logging import init_logging, set_logging_level
import matplotlib.pyplot as plt
import seaborn as sns
from matplotlib.colors import LogNorm
import matplotlib.pyplot as plt



def collect_times(root_path: str, column_nums: list,
    row_col_nums: list, exp_names: list, skip_dbs: list,
    exp_paths: dict, num_measurement: int):
    XLSX_NAME = "time.xlsx"
    df_measurement_exps = {}

    for exp_name in exp_names:
        #Initialize paths for the current experiment
        exp_path = exp_paths[exp_name]
        perf_path = os.path.join(root_path, exp_path)
        df_measurement = pd.DataFrame(columns=column_nums)
        gather_times_path = os.path.join(perf_path, XLSX_NAME)

        out_workbook = Workbook()
        out_workbook.remove(out_workbook.active)
        out_sheet = out_workbook.create_sheet("Times")

        for db_idx, db_num in enumerate(range(1, 21)):
            db_name = "DBCartesianProduct{}".format(db_num)
            if db_num in skip_dbs[exp_name]:
                continue
            path_xlsx = os.path.join(perf_path, db_name, XLSX_NAME)
            workbook = load_workbook(filename=path_xlsx)
            sheet = workbook.active
            row_count = sheet.max_row

            col_idx = db_idx + 1
            col_letter = get_column_letter(col_idx)

            # Writng header files
            out_sheet.cell(row=1, column=col_idx).value = db_name

            # Collecting average.
            start_row_idx = 2
            np_measures = np.zeros(num_measurement)
            for row_idx_dst in range(start_row_idx, start_row_idx + num_measurement):
                offset_idx = row_idx_dst - start_row_idx
                row_idx_src = row_count - num_measurement  + offset_idx
                val = sheet.cell(row=row_idx_src, column=2).value
                np_measures[offset_idx] = float(val)
                out_sheet.cell(row=row_idx_dst, column=col_idx).value = val

            time_avg = np.mean(np_measures[1:])
            row_num = row_col_nums[db_idx][0]
            col_num = row_col_nums[db_idx][1]
            df_measurement.at[row_num, col_num] = time_avg

        out_workbook.save(gather_times_path)
        df_measurement = df_measurement.astype(float)
        df_measurement_exps[exp_name] = df_measurement

    return df_measurement_exps


def dump_results_to_latex(figaro_impls: list, df_measurement_exps: dict):
    root_dir = "results/synthetic"
    out_path = os.path.join(root_dir, "synthetic_perf.tex")
    os.makedirs(root_dir, exist_ok=True)
    out_file = open(out_path, "w")

    for figaro_impl in figaro_impls:
        df_figaro_thin =  df_measurement_exps[figaro_impl]
        out_file.write("{} performance\n".format(figaro_impl))
        df_figaro_thin_lat = df_figaro_thin.rename(index={512:" $2^9$", 1024: "$2^{10}$",
        2048: "$2^{11}$", 4096: "$2^{12}$", 8192: "$2^{13}$"},
        columns={64:"2^6", 128:"$2^7$", 256: "$2^8$", 512: "$2^9$", 1024:"$2^{10}$", 4096: "$2^{12}$"})
        out_file.write(df_figaro_thin_lat.to_latex(float_format="%.2f",escape=False))
        out_file.write("\n")

        df_measurement_speed_up = df_measurement_exps["postprocess_mkl"] / df_measurement_exps[figaro_impl]
        out_file.write("Speed up {}\n".format(figaro_impl))
        df_measurement_speed_up_lat = df_measurement_speed_up.rename(index={512:" $2^9$", 1024: "$2^{10}$",
        2048: "$2^{11}$", 4096: "$2^{12}$", 8192: "$2^{13}$"},
        columns={64:"2^6", 128:"$2^7$", 256: "$2^8$", 512: "$2^9$", 1024:"$2^{10}$", 4096: "$2^{12}$"})
        out_file.write(df_measurement_speed_up_lat.to_latex(float_format="%.0f",escape=False, na_rep=" "))
        out_file.write("\n")

    out_file.close()



def main(args):
    parser = argparse.ArgumentParser()
    parser.add_argument("--root_path", action="store",
                        dest="root_path", required=True)
    parser.add_argument("--dump_results", action="store_true",
        dest="dump_results", required=False)
    args = parser.parse_args(args)


    root_path = args.root_path
    dump_results = args.dump_results

    num_measurement = 21


    figaro_impls = ["figaro_thin", "figaro_lapack"]
    exp_names = [*figaro_impls, "postprocess_mkl"]
    exp_paths = {"figaro_thin": "comparisons/performance/figaro/only_r/thin_diag/thread48",
                "figaro_lapack": "comparisons/performance/figaro/only_r/lapack/thread48",
                "mkl": "comparisons/performance/python/mkl",
                "postprocess_mkl": "comparisons/performance/postprocess/lapack/col_major/only_r/thread48",}

    column_nums = [64, 256, 1024, 4096]
    row_nums = [512, 1024, 2048, 4096, 8192]
    row_col_nums = []

    skip_dbs = {"mkl": [12, 15, 16, 18, 19, 20], "postprocess_mkl": [12, 15, 16, 18, 19, 20],"figaro_thin": [],  "figaro_lapack":[]}

    for row_num in row_nums:
        for col_num in column_nums:
            row_col_nums.append((row_num, col_num))

    df_measurement_exps = collect_times(root_path, column_nums,
            row_col_nums, exp_names, skip_dbs, exp_paths, num_measurement)
    if dump_results:
        dump_results_to_latex(figaro_impls, df_measurement_exps)


if __name__ == "__main__":
    init_logging()
    set_logging_level(logging.INFO)
    main(sys.argv[1:])