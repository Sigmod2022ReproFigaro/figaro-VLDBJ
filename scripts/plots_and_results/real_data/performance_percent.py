""""
Figaro times collection times script.
"""
import csv
import numpy as np
import pandas as pd
import argparse
import os
from openpyxl import load_workbook
from openpyxl.utils.cell import get_column_letter
import logging
import argparse
import sys
import numpy as np
from evaluation.custom_logging import init_logging, set_logging_level
import matplotlib.pyplot as plt


def collect_times(ohe: bool, root_path: str, exp_names: list, db_names: list,
     exp_paths: dict, join_orders: dict, start_per: int, end_per: int,
         per_inc: int, num_measurement: int):
    XLSX_NAME = "time.xlsx"
    df_measurement_exps = {}
    for exp_name in exp_names:
        exp_path = exp_paths[exp_name]
        perf_path = os.path.join(root_path, exp_path)
        df_measurement = pd.DataFrame(columns=db_names)

        for db_name in db_names:
            for db_idx, percent in enumerate(range(start_per, end_per + 1, per_inc)):
                if not ohe:
                    db_name_per = "{}{}".format(db_name, percent)
                else:
                    db_name_per = "{}PK1C{}".format(db_name, percent)
                join_order = join_orders[db_name]
                path_xlsx = os.path.join(perf_path, db_name_per, join_order, XLSX_NAME)
                workbook = load_workbook(filename=path_xlsx, data_only=True)
                sheet = workbook.active
                row_count = sheet.max_row
                col_idx = db_idx + 1
                col_letter = get_column_letter(col_idx)

                start_row_idx = 2
                np_measures = np.zeros(num_measurement)
                for row_idx_dst in range(start_row_idx, start_row_idx + num_measurement):
                    offset_idx = row_idx_dst - start_row_idx
                    row_idx_src = row_count - num_measurement  + offset_idx
                    val = sheet.cell(row=row_idx_src, column=2).value
                    np_measures[offset_idx] = float(val)

                time_avg = np.mean(np_measures[1:])
                df_measurement.at[percent, db_name] = time_avg

            df_measurement.index = df_measurement.index.set_names(['index'])
            df_measurement = df_measurement.astype(float)
        df_measurement_exps[exp_name] = df_measurement

    return df_measurement_exps


def dump_results_to_dat(ohe: bool, db_names: list, df_measurement_exps: dict):
    db_name_map = {
    "DBFavorita": "exp1perf-favorita.dat",
    "DBRetailer": "exp1perf-retailer.dat",
    "DBYelp": "exp1perf-yelp.dat"
    }
    exp_dat_names = ["#percentage of data", "figaro-time", "mkl-time", "postproc-thin-time"]
    exp_to_show = ["figaro_thin", "post_proc_mkl", "post_proc_thin"]

    for db_name in db_names:
        dbs_results = []
        for exp_name in exp_to_show:
            dbs_results.append(df_measurement_exps[exp_name][db_name])
        df_db_results = pd.concat(dbs_results, axis=1)
        df_db_results = df_db_results.reset_index().rename(columns={df_db_results.index.name:'index'})
        df_db_results.columns = exp_dat_names
        dir_out_root = "results"
        dir_out = "ohe" if ohe else "non-ohe"
        dir_out = os.path.join(dir_out_root, dir_out)
        os.makedirs(dir_out, exist_ok=True)
        out_name = os.path.join(dir_out, db_name_map[db_name])
        df_db_results.to_csv(out_name, float_format='%.2f', sep='\t', index=False, quoting=csv.QUOTE_NONE,  escapechar=" ")

def plot_performance(ohe: bool, exp_names: list,
        df_measurement_exps: dict):
    plt.figure("name", figsize=(16, 8), dpi=80)
    plt.xlabel("Percentage of input matrix")
    plt.ylabel("wall-clock-time[s]")

    plt.title("Runtime performance comparison of Figaro and competitors on real-world datasets")

    plt.yscale('linear')
    plt.locator_params(axis='x', nbins=6)
    plt.locator_params(axis='x', nbins=10)
    db_marker =  {"DBFavorita": "^", "DBYelp": "s", "DBRetailer": "x"}
    exp_colour = {"figaro_thin": "r", "mkl": "b", "openblas": "g",
    "post_proc_thin": "y", "post_proc_mkl": "c", "figaro_lapack": "m"
    }

    for exp_name in exp_names:
        df_measurement = df_measurement_exps[exp_name]
        for db_name in df_measurement:
            plt.plot(df_measurement[db_name], "-" + exp_colour[exp_name] + db_marker[db_name], label="{} {}".format(exp_name, db_name), markersize=8)
    plt.legend(loc="upper left")

    plt.show()
    dir_out_root = "results"
    dir_out = "ohe" if ohe else "non-ohe"
    dir_out = os.path.join(dir_out_root, dir_out)
    os.makedirs(dir_out, exist_ok=True)
    out_path = os.path.join(dir_out, "performance_percent.png")
    plt.savefig(out_path)


def main(args):
    parser = argparse.ArgumentParser()
    parser.add_argument("--root_path", action="store",
                        dest="root_path", required=True)
    parser.add_argument("-e", "--exp_names", dest="exp_names", nargs='*', required=True)
    parser.add_argument("--dump_results", action="store_true",
        dest="dump_results", required=False)
    parser.add_argument("--ohe", action="store_true",
        dest="ohe", required=False)
    args = parser.parse_args(args)


    root_path = args.root_path
    exp_names = args.exp_names
    dump_results = args.dump_results
    ohe = args.ohe
    #exp_names = ["figaro_thin",  "post_proc_mkl", "post_proc_thin"]

    exp_paths = {"figaro_thin": "comparisons/performance/figaro/only_r/thin_diag/thread48", "mkl": "comparisons/performance/python/mkl",
    "figaro_lapack": "comparisons/performance/figaro/only_r/lapack/thread48",
    "openblas": "comparisons/performance/python/openblas",
    "post_proc_thin": "comparisons/performance/postprocess/thin_diag/thread48",
    "post_proc_mkl": "comparisons/performance/postprocess/lapack/col_major/only_r/thread48"}

    db_names = ["DBRetailer", "DBFavorita", "DBYelp"]
    if ohe:
        join_orders = {"DBRetailer": "ItemRoot",
        "DBFavorita": "ItemsRoot", "DBYelp": "UserRoot"}
    else:
        join_orders = {"DBRetailer": "LocationRoot",
            "DBFavorita": "StoresRoot", "DBYelp": "BusinessRoot"}

    start_per = 10
    end_per = 100
    per_inc = 10

    num_measurement = 21

    df_measurement_exps = collect_times(ohe, root_path, exp_names, db_names,
        exp_paths, join_orders, start_per, end_per, per_inc,
        num_measurement)
    if dump_results:
        dump_results_to_dat(ohe, db_names, df_measurement_exps)
    plot_performance(ohe, exp_names, df_measurement_exps)

if __name__ == "__main__":
    init_logging()
    set_logging_level(logging.INFO)
    main(sys.argv[1:])