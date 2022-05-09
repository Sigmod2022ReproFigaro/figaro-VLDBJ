""""
Figaro times collection times script.
"""
import numpy as np
import pandas as pd
import os
from openpyxl import load_workbook, Workbook
from openpyxl.utils.cell import get_column_letter
import argparse
import matplotlib.pyplot as plt
from matplotlib import ticker
import csv
import logging
import argparse
import sys
import numpy as np
from evaluation.custom_logging import init_logging, set_logging_level
import matplotlib.pyplot as plt


def collect_times(ohe: bool, root_path: str, exp_name: str, db_names: list,
        exp_paths: dict, join_orders: dict, thread_nums_exp: dict,
            num_measurement: int):
    XLSX_NAME = "time.xlsx"
    df_measurement_exp_dbs = {}
    exp_path = exp_paths[exp_name]
    perf_path = os.path.join(root_path, exp_path)

    for db_name in db_names:
        gather_times_path = os.path.join(perf_path, db_name + XLSX_NAME)
        out_workbook = Workbook()
        out_workbook.remove(out_workbook.active)
        out_sheet = out_workbook.create_sheet("Times")
        df_measurement = pd.DataFrame(columns=join_orders[db_name])
        for join_idx, join_order in enumerate(join_orders[db_name]):
            for thread_idx, thread_num in enumerate(thread_nums_exp[exp_name]):
                if ohe:
                    db_name_pref =  "{}PK1C100".format(db_name)
                else:
                    db_name_pref = "{}100".format(db_name)
                db_name_full = "{} {} {}".format(db_name_pref, join_order, thread_num)
                path_xlsx = os.path.join(perf_path, "thread"+str(thread_num), db_name_pref, join_order, XLSX_NAME)
                if os.path.isfile(path_xlsx):
                    workbook = load_workbook(filename=path_xlsx, data_only=True)
                    sheet = workbook.active
                    row_count = sheet.max_row
                    col_idx = join_idx * len(thread_nums_exp[exp_name]) + thread_idx + 1
                    col_letter = get_column_letter(col_idx)

                    out_sheet.cell(row=1, column=col_idx).value = db_name_full

                    start_row_idx = 2
                    np_measures = np.zeros(num_measurement)
                    for row_idx_dst in range(start_row_idx, start_row_idx + num_measurement):
                        offset_idx = row_idx_dst - start_row_idx
                        row_idx_src = row_count - num_measurement  + offset_idx
                        val = sheet.cell(row=row_idx_src, column=2).value
                        np_measures[offset_idx] = float(val)
                        out_sheet.cell(row=row_idx_dst, column=col_idx).value = val

                    out_sheet.cell(row=start_row_idx + num_measurement, column=col_idx).value = '=AVERAGE({}{}:{}{})'.format(col_letter, start_row_idx + 1,
                            col_letter,22)
                    time_avg = np.mean(np_measures[1:])
                else:
                    time_avg = float("inf")
                df_measurement.at[thread_num, join_order] = time_avg

        out_workbook.save(gather_times_path)
        df_measurement_exp_dbs[(exp_name, db_name)] = df_measurement.astype(np.float64)

    return df_measurement_exp_dbs


def dump_results_to_dat(ohe: bool, db_names: list, df_measurement_exps: dict,
    thread_nums_exp: dict):
    exp_name = "figaro_thin"
    exp_dat_name = "exp2cores.dat"
    exp_dat_names = ["#cores", "Retailer", "Favorita", "Yelp"]
    dbs_results = []

    db_name_map = {
        "DBFavorita": "exp3perf-favorita.dat",
        "DBRetailer": "exp3perf-retailer.dat",
        "DBYelp": "exp3perf-yelp.dat"
    }

    for db_name in db_names:
        df_measurement = df_measurement_exps[(exp_name, db_name)]
        # Dumping join order times to a .dat file
        dir_out_root = "results"
        dir_out = "ohe" if ohe else "non-ohe"
        dir_out = os.path.join(dir_out_root, dir_out)
        os.makedirs(dir_out, exist_ok=True)
        out_name = os.path.join(dir_out, db_name_map[db_name])

        df_measurement.to_csv(out_name, float_format='%.2f', sep='\t', index=False, quoting=csv.QUOTE_NONE,  escapechar=" ")

        min_idx_cols = df_measurement.idxmin(axis="columns")
        max_num_threads = thread_nums_exp[exp_name][-1]
        min_join_order = min_idx_cols[max_num_threads]
        dbs_results.append(df_measurement[min_join_order])

    df_db_results = pd.concat(dbs_results, axis=1)
    df_db_results = df_db_results.reset_index().rename(columns={df_db_results.index.name:'index'})
    df_db_results.columns = exp_dat_names
    dir_out_root = "results"
    dir_out = "ohe" if ohe else "non-ohe"
    dir_out = os.path.join(dir_out_root, dir_out)
    os.makedirs(dir_out, exist_ok=True)
    out_name = os.path.join(dir_out, exp_dat_name)
    df_db_results.to_csv(out_name, float_format='%.2f', sep='\t', index=False, quoting=csv.QUOTE_NONE,  escapechar=" ")


def plot_performance(ohe: bool, db_names: list, df_measurement_exp_dbs: dict, thread_nums_exp: dict):
    plt.figure("name", figsize=(16, 8), dpi=80)
    plt.xlabel("Number of threads")
    plt.ylabel("wall-clock-time[s]")

    plt.title("Runtime performance comparison of Figaro and competitors on real-world datasets")


    plt.yscale('log', base=10)
    plt.xscale('log', base=2)
    #plt.locator_params(axis='x', nbins=6)
    #plt.locator_params(axis='y', nbins=20)
    db_marker =  {"DBFavorita": "^", "DBYelp": "s", "DBRetailer": "x"}
    db_colour =  {"DBFavorita": "r", "DBYelp": "c", "DBRetailer": "g"}

    exp_name = "figaro_thin"
    for db_name in db_names:
        df_measurement = df_measurement_exp_dbs[(exp_name, db_name)]
        min_idx_cols = df_measurement.idxmin(axis="columns")
        max_num_threads = thread_nums_exp[exp_name][-1]
        min_join_order = min_idx_cols[max_num_threads]

        plt.plot(df_measurement[min_join_order], "-" + db_colour[db_name] + db_marker[db_name], label="{} {} {}".format(exp_name, db_name, min_join_order), markersize=10)

    ax = plt.gca()
    def myLogFormat(y,pos):
        # Find the number of decimal places required
        decimalplaces = int(np.maximum(-np.log10(y),0))     # =0 for numbers >=1
        # Insert that number into a format string
        formatstring = '{{:.{:1d}f}}'.format(decimalplaces)
        # Return the formatted tick label
        return formatstring.format(y)

    ax.yaxis.set_major_formatter(ticker.FuncFormatter(myLogFormat))

    #formatter = ticker.LogFormatter(base=10, labelOnlyBase=False)
    #formatter.set_scientific(True)
    #formatter.set_powerlimits((-1,1))
    #ax.yaxis.set_major_formatter(formatter)

    formatter = ticker.LogFormatter(base=2)
    #formatter.set_scientific(True)
    #formatter.set_powerlimits((0,48))
    ax.xaxis.set_major_formatter(formatter)
    plt.legend(loc="upper right")
    dir_out_root = "results"
    dir_out = "ohe" if ohe else "non-ohe"
    dir_out = os.path.join(dir_out_root, dir_out)
    os.makedirs(dir_out, exist_ok=True)
    out_path = os.path.join(dir_out, "exp2threads.png")
    plt.savefig(out_path)

    plt.show()

def main(args):
    parser = argparse.ArgumentParser()
    parser.add_argument("--root_path", action="store",
                        dest="root_path", required=True)
    parser.add_argument("-e", "--exp_name", dest="exp_name",  required=True)
    parser.add_argument("--dump_results", action="store_true",
        dest="dump_results", required=False)
    parser.add_argument("--ohe", action="store_true",
        dest="ohe", required=False)
    args = parser.parse_args(args)


    root_path = args.root_path
    exp_name = args.exp_name
    dump_results = args.dump_results
    ohe = args.ohe

    db_names = ["DBRetailer", "DBFavorita", "DBYelp"]
    #db_names = ["DBRetailerPK1C100", "DBFavoritaPK1C100", "DBYelpPK1C100"]
    figaro_threads = [1, 2, 4, 8, 16, 24, 32, 48]
    thread_nums_exp = {"mkl": [48], "figaro_thin": figaro_threads}

    join_orders = {"DBFavorita": ["HolidaysRoot", "ItemsRoot", "OilRoot", "SalesRoot", "StoresRoot", "TransactionsRoot"],
                "DBYelp": ["BusinessRoot", "CategoryRoot", "CheckinRoot", "HoursRoot", "ReviewRoot", "UserRoot"],
                "DBRetailer": ["CensusRoot", "InventoryRoot", "ItemRoot", "LocationRoot", "WeatherRoot"]}

    exp_paths = {"figaro_thin": "comparisons/performance/figaro/only_r/thin_diag",
    "mkl": "comparisons/performance/python/mkl",
    "figaro_lapack": "comparisons/performance/figaro/only_r/lapack",
    "openblas": "comparisons/performance/python/openblas",
    "post_proc_thin": "comparisons/performance/postprocess/thin_diag",
    "post_proc_mkl": "comparisons/performance/postprocess/lapack"}

    db_names = ["DBRetailer", "DBFavorita", "DBYelp"]

    num_measurement = 21

    df_measurement_exps = collect_times(ohe, root_path, exp_name, db_names,
        exp_paths, join_orders, thread_nums_exp,
        num_measurement)
    if dump_results:
        dump_results_to_dat(ohe, db_names, df_measurement_exps, thread_nums_exp)
    plot_performance(ohe, db_names,  df_measurement_exps,  thread_nums_exp)

if __name__ == "__main__":
    init_logging()
    set_logging_level(logging.INFO)
    main(sys.argv[1:])