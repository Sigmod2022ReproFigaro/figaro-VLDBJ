import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill
import logging

class AccuracyWorkbook:

    def __init__(self, output_file, precision, operation):
        self.work_book = Workbook()
        self.work_book.remove_sheet(self.work_book.active)
        self.prec_ws = self.work_book.create_sheet("precision")
        self.figaro_ws = self.work_book.create_sheet("Figaro")
        self.comp_ws = self.work_book.create_sheet(operation.upper())
        self.prec_rel_figaro_ws = self.work_book.create_sheet("Figaro_rel")
        self.prec_rel_op_ws = self.work_book.create_sheet(operation.upper()+"_rel")
        self.precision = precision
        self.output_file = output_file
        self.operation = operation.upper()
        logging.info("Precision is {}".format(self.precision))

    def save_entry(self, row, col, figaro_val, compet_val):
        diff = abs(figaro_val - compet_val)
        diff_rel_figaro = 0 if figaro_val == 0 else abs(diff / figaro_val)
        diff_rel_op = 0 if compet_val == 0 else abs(diff / compet_val)

        '''
        logging.info("""Differrence row: {} col: {}
                        figaro_val {} compet_val {}
                        diff {} diff_rel_figaro {} diff_rel_op {}""".
                    format(row, col, figaro_val, compet_val, diff,
                    diff_rel_figaro, diff_rel_op))
        '''
        self.figaro_ws.cell(row, col).value = figaro_val
        self.comp_ws.cell(row, col).value = compet_val
        self.prec_ws.cell(row, col).value = diff
        self.prec_rel_figaro_ws.cell(row, col).value = diff_rel_figaro
        self.prec_rel_op_ws.cell(row, col).value = diff_rel_op

        red = openpyxl.styles.colors.Color(rgb='00FF0000')
        red_fill = PatternFill(patternType='solid', fgColor=red)
        green = openpyxl.styles.colors.Color(rgb='00008000')
        green_fill = PatternFill(patternType='solid', fgColor=green)

        self.prec_ws.cell(row, col).fill = red_fill if diff >= self.precision else green_fill
        self.prec_rel_op_ws.cell(row, col).fill = red_fill if diff_rel_op >= self.precision else green_fill
        self.prec_rel_figaro_ws.cell(row, col).fill = red_fill if diff_rel_figaro >= self.precision else green_fill


    def save(self):
        self.work_book.save(self.output_file)