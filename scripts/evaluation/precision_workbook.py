from argparse import ArgumentParser
from decimal import Decimal
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill


class PrecisionWorkbook:

    def __init__(self, output_file, precision, operation):
        self.work_book = Workbook()
        self.work_book.remove_sheet(self.work_book.active)
        self.prec_ws = self.work_book.create_sheet("precision")
        self.lmfao_ws = self.work_book.create_sheet("LMFAO")
        self.comp_ws = self.work_book.create_sheet(operation.upper())
        self.prec_rel_lmfao_ws = self.work_book.create_sheet("LMFAO_rel")
        self.prec_rel_op_ws = self.work_book.create_sheet(operation.upper()+"_rel")
        self.precision = precision
        self.output_file = output_file
        self.operation = operation.upper()

    def save_entry(self, row, col, lmfao_val, comp_val):
        diff = abs(lmfao_val - comp_val)
        diff_rel_lmfao = 0 if lmfao_val == 0 else abs(diff / lmfao_val)
        diff_rel_op = 0 if comp_val == 0 else abs(diff / comp_val)

        self.lmfao_ws.cell(row, col).value = lmfao_val
        self.comp_ws.cell(row, col).value = comp_val
        self.prec_ws.cell(row, col).value = diff
        self.prec_rel_lmfao_ws.cell(row, col).value = diff_rel_lmfao
        self.prec_rel_op_ws.cell(row, col).value = diff_rel_op

        red = openpyxl.styles.colors.Color(rgb='00FF0000')
        red_fill = PatternFill(patternType='solid', fgColor=red)
        green = openpyxl.styles.colors.Color(rgb='00008000')
        green_fill = PatternFill(patternType='solid', fgColor=green)

        self.prec_ws.cell(row, col).fill = red_fill if diff >= self.precision else green_fill
        self.prec_rel_op_ws.cell(row, col).fill = red_fill if diff_rel_op >= self.precision else green_fill
        self.prec_rel_lmfao_ws.cell(row, col).fill = red_fill if diff_rel_lmfao >= self.precision else green_fill

    def save(self):
        self.work_book.save(self.output_file)