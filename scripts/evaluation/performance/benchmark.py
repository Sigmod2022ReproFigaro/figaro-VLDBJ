from argparse import ArgumentParser
from pathlib import Path
from openpyxl.utils.cell import get_column_letter
import re
from openpyxl import Workbook, load_workbook
import logging

reg_pattern = "(?:[. ]*?)(?<=##Figaro####)([\w ]+)(?:##)(?:[ ]*)([-+]?[0-9]+\.?[0-9]*([eE]?[-+]?[0-9]*))"

class TimeWorkbook:

    def __init__(self, time_file_for, data_op):
        path_time_file_for = Path(time_file_for)
        if path_time_file_for.is_file():
            self.work_book = load_workbook(path_time_file_for)
            if data_op in self.work_book.sheetnames:
                logging.info("File {} exists. Loading.".format(path_time_file_for))
                self.time_sheet = self.work_book[data_op]
            else:
                self.time_sheet = self.work_book.create_sheet(data_op)
        else:
            self.work_book = Workbook()
            self.work_book.remove_sheet(self.work_book.active)
            self.time_sheet = self.work_book.create_sheet(data_op)
        self.output_file = time_file_for
        self.col_header = 1
        self.row_header = 1


    def get_db_idx(self, db_name: str)-> int:
        for col_idx in range(1, self.time_sheet.max_column + 1):
            if db_name == self.time_sheet.cell(self.row_header, col_idx).value:
                return col_idx

        col_idx = self.time_sheet.max_column + 1
        return col_idx


    def save_entry(self, row, col, entry):
        self.time_sheet.cell(row, col).value = entry


    def save_entries(self, db_name: str, measurement_times):
        db_name_idx = self.get_db_idx(db_name)
        self.save_entry(self.row_header, self.col_header, "Measure/dataset")
        self.save_entry(self.row_header, db_name_idx, db_name)

        row_idx = self.row_header + 1
        col_idx = db_name_idx
        logging.info("DB Idx {}".format(db_name_idx))
        for measurement in measurement_times:
            time = measurement_times[measurement]
            n_times = len(time)
            self.save_entry(row_idx + len(time), self.col_header,  measurement)
            for idx, timeSpec in enumerate(time, 1):
                self.save_entry(row_idx, self.col_header, measurement + str(idx))
                self.save_entry(row_idx, col_idx, timeSpec)
                row_idx += 1
            col_letter = get_column_letter(col_idx)
            if n_times == 1:
                self.save_entry(row_idx, col_idx, time[0])
            else:
                self.save_entry(row_idx, col_idx,
                '=AVERAGE({}{}:{}{})'.format(
                    col_letter,row_idx - n_times + 1,
                    col_letter,row_idx - 1))

            row_idx += 1

    def save(self):
        self.work_book.save(self.output_file)


def parse_times(times_path):
    meas_times = {}
    with open(times_path) as time_file:
        for line in time_file:
            print(line)
            matches = re.findall(reg_pattern, line)
            if len(matches) > 0:
                print("Match {}".format(line))
                print(matches)
                measurement = matches[0][0]
                time = float(matches[0][1])
                #print(measurement)
                #print(time)

                if measurement in meas_times:
                    meas_times[measurement].append(time)
                else:
                    meas_times[measurement] = [time]
    #print(meas_times)
    return meas_times

#TODO: Get DB name automatically by reading openpyxl
#TODO: ADd option for save_entries for average or microbench

def gather_times(log_path: str, times_formated_path: str,
                 database_name: str):
    measurement_times = parse_times(log_path)

    time_workbook = TimeWorkbook(times_formated_path, 'qr')
    time_workbook.save_entries(database_name, measurement_times)
    time_workbook.save()


#TODO: Collect all times into one big sheet
#TODO: Extract all sizes and add this to one big sheet

if __name__ == "__main__":
    parser = ArgumentParser()
    parser.add_argument("-t", "--time_file_for", dest="time_file_for", required=True,
                        help='.xlsx file in which time measurements are stored')
    parser.add_argument("-i", "--times_path", dest="times_path", required=True,
                        help='log file where times are logged while runtime')
    parser.add_argument("-ds", "--data_set", dest="data_set", required=True)
    #parser.add_argument("-op", "--data_op", dest="data_op", required=True)
    #parser.add_argument("-s", "--data_set_idx", dest="data_set_idx", required=True)
    args = parser.parse_args()
    data_set = args.data_set
    time_file_for = args.time_file_for
    times_path = args.times_path
    #data_op = args.data_op
    #data_set_idx = int(args.data_set_idx) + 1

    #time_workbook = TimeWorkbook(time_file_for, data_op)

    data_set_times = {}
    measure_times = parse_times(times_path, time_file_for)
    #time_workbook.save_entries(data_set, measure_times, data_set_idx)
    #time_workbook.save()

